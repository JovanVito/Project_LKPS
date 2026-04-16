import json
from django.conf import settings
import google.generativeai as genai
from django.views.decorators.csrf import csrf_exempt
from django.db import IntegrityError 
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib import messages
from django.db import transaction
from django.http import JsonResponse, HttpResponse, Http404
from django.contrib.auth.decorators import user_passes_test
from django.db import connection
from docxtpl import DocxTemplate
from docxtpl import InlineImage
from docx.shared import Mm
from .models import (
    IdentitasPengusul, TimPenyusun, ProgramStudi, ProfilPengguna,
    Tabel_1A1, Tabel_1A2_Sumber, Tabel_1A3_Penggunaan, Tabel_1A4, Tabel_1A5, Tabel_1B_SPMI, 
    Tabel_2A_Mahasiswa, Tabel_2A2_Asal, Tabel_2A3_Kondisi, 
    Tabel_2B1_MK, Tabel_2B2_CPL, Tabel_2B3_Pemenuhan, Tabel_2B4_MasaTunggu, Tabel_2B5_BidangKerja, Tabel_2B6_Kepuasan, Tabel_2B_Summary,
    Tabel_2C_Fleksibilitas, Tabel_2D_Rekognisi,
    Tabel_3_Summary,Tabel_3A1_Sarana,Tabel_3A2_Penelitian, Tabel_3A3_Pengembangan_DTPR, Tabel_3C1_Kerjasama,Tabel_3C2_Publikasi,Tabel_3C3_HKI, 
    Tabel_4A1_Sarana, Tabel_4A2_PkM, Tabel_4C1_Kerjasama, Tabel_4C2_Diseminasi, Tabel_4C3_HKI, Tabel_4_Summary, 
    Tabel_5_1_TataKelola, Tabel_5_2_Sarana, Tabel_6_Misi
)

# ==================================================================
# MASTER MAPPING — Satu Dictionary untuk Semua Tabel Sederhana
# ==================================================================
MASTER_MAPPING = {
    '1a1': {
        'model': Tabel_1A1,
        'template': 'lkps_app/tabel_1a1.html',
        'context_names': {'data': 'data'},
        'mapping': {
            'unit_kerja[]': {'field': 'unit_kerja', 'type': 'str'},
            'nama_ketua[]': {'field': 'nama_ketua', 'type': 'str'},
            'periode_jabatan[]': {'field': 'periode_jabatan', 'type': 'str'},
            'pendidikan_terakhir[]': {'field': 'pendidikan_terakhir', 'type': 'str'},
            'jabatan_fungsional[]': {'field': 'jabatan_fungsional', 'type': 'str'},
            'tupoksi[]': {'field': 'tupoksi', 'type': 'str'},
        },
    },
    '1b': {
        'model': Tabel_1B_SPMI,
        'template': 'lkps_app/tabel_1b.html',
        'context_names': {'data_1b': 'data_1b'},
        'mapping': {
            'nama_unit[]': {'field': 'nama_unit', 'type': 'str'},
            'dokumen[]': {'field': 'dokumen', 'type': 'str'},
            'jml_auditor_cert[]': {'field': 'jml_auditor_cert', 'type': 'int'},
            'jml_auditor_non[]': {'field': 'jml_auditor_non', 'type': 'int'},
            'frekuensi_audit[]': {'field': 'frekuensi_audit', 'type': 'int'},
            'link_pt[]': {'field': 'link_pt', 'type': 'str'},
            'link_upps[]': {'field': 'link_upps', 'type': 'str'},
        },
    },
    '2b1': {
        'model': Tabel_2B1_MK,
        'template': 'lkps_app/tabel_2b1.html',
        'context_names': {'data_mk': 'data_mk'},
        'mapping': {
            'nama_mk[]': {'field': 'nama_mk', 'type': 'str'},
            'sks_mk[]': {'field': 'sks', 'type': 'int'},
            'smt_mk[]': {'field': 'semester', 'type': 'int'},
            'pl1[]': {'field': 'pl1', 'type': 'bool'},
            'pl2[]': {'field': 'pl2', 'type': 'bool'},
            'pl3[]': {'field': 'pl3', 'type': 'bool'},
            'pl4[]': {'field': 'pl4', 'type': 'bool'},
        },
    },
    '2b2': {
        'model': Tabel_2B2_CPL,
        'template': 'lkps_app/tabel_2b2.html',
        'context_names': {'data_cpl': 'data_cpl'},
        'mapping': {
            'kode_cpl[]': {'field': 'kode_cpl', 'type': 'str'},
            'map_pl1[]': {'field': 'pl1', 'type': 'bool'},
            'map_pl2[]': {'field': 'pl2', 'type': 'bool'},
            'map_pl3[]': {'field': 'pl3', 'type': 'bool'},
            'map_pl4[]': {'field': 'pl4', 'type': 'bool'},
        },
    },
    '2b3': {
        'model': Tabel_2B3_Pemenuhan,
        'template': 'lkps_app/tabel_2b3.html',
        'context_names': {'data_pemenuhan': 'data_pemenuhan'},
        'mapping': {
            'pem_cpl[]': {'field': 'cpl', 'type': 'str'},
            'pem_cpmk[]': {'field': 'cpmk', 'type': 'str'},
            'pem_smt1[]': {'field': 'smt1', 'type': 'str'},
            'pem_smt2[]': {'field': 'smt2', 'type': 'str'},
            'pem_smt3[]': {'field': 'smt3', 'type': 'str'},
        },
    },
    '2c': {
        'model': Tabel_2C_Fleksibilitas,
        'template': 'lkps_app/tabel_2c.html',
        'context_names': {'data_2c': 'data_2c'},
        'mapping': {
            'bentuk_pembelajaran[]': {'field': 'bentuk_pembelajaran', 'type': 'str'},
            'ts_2[]': {'field': 'ts_2', 'type': 'int'},
            'ts_1[]': {'field': 'ts_1', 'type': 'int'},
            'ts[]': {'field': 'ts', 'type': 'int'},
            'link_bukti[]': {'field': 'link_bukti', 'type': 'str'},
        },
    },
    '2d': {
        'model': Tabel_2D_Rekognisi,
        'template': 'lkps_app/tabel_2d.html',
        'context_names': {'data_2d': 'data_2d'},
        'mapping': {
            'sumber[]': {'field': 'sumber', 'type': 'str'},
            'jenis_pengakuan[]': {'field': 'jenis_pengakuan', 'type': 'str'},
            'ts_2[]': {'field': 'ts_2', 'type': 'int'},
            'ts_1[]': {'field': 'ts_1', 'type': 'int'},
            'ts[]': {'field': 'ts', 'type': 'int'},
            'link_bukti[]': {'field': 'link_bukti', 'type': 'str'},
        },
    },
    '3a3': {
        'model': Tabel_3A3_Pengembangan_DTPR,
        'template': 'lkps_app/tabel_3a3.html',
        'context_names': {'data_3a3': 'data_3a3'},
        'mapping': {
            'jenis_pengembangan[]': {'field': 'jenis_pengembangan', 'type': 'str'},
            'nama_dosen[]': {'field': 'nama_dosen', 'type': 'str'},
            'ts_2[]': {'field': 'ts_2', 'type': 'int'},
            'ts_1[]': {'field': 'ts_1', 'type': 'int'},
            'ts[]': {'field': 'ts', 'type': 'int'},
            'link_bukti[]': {'field': 'link_bukti', 'type': 'str'},
        },
    },
}

# ==================================================================
# DYNAMIC VIEW — Satu Fungsi untuk Semua Tabel di MASTER_MAPPING
# ==================================================================
def halaman_tabel_dinamis(request, kode_tabel):
    """
    View generik yang menangani GET (render halaman) dan POST (autosave AJAX)
    untuk semua tabel sederhana yang terdaftar di MASTER_MAPPING.
    """
    konfig = MASTER_MAPPING.get(kode_tabel)
    if not konfig:
        raise Http404("Tabel tidak ditemukan")

    # Handle AJAX autosave POST
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return universal_table_autosave(request, konfig['model'], konfig['mapping'])

    # Handle GET — render halaman dengan context yang sesuai template HTML lama
    context_data = {}
    for ctx_key in konfig.get('context_names', {}):
        context_data[ctx_key] = konfig['model'].objects.all().order_by('id')

    return render(request, konfig['template'], context_data)


# ==========================================
# EXPORT TO WORD
# ==========================================
def export_lkps_word(request):
    try:
        template_path = str(settings.BASE_DIR / 'static' / 'lkps_app' / 'word_templates' / 'Master_Format_LKPS.docx')
        doc = DocxTemplate(template_path)
        identitas_data = IdentitasPengusul.objects.first()

        context = {
            'identitas': identitas_data or {},
            'tim_penyusun': TimPenyusun.objects.all(),
            't1a1': Tabel_1A1.objects.all(),
            't1a2': Tabel_1A2_Sumber.objects.all(),
            't1a3': Tabel_1A3_Penggunaan.objects.all(),
            't1a4': Tabel_1A4.objects.all(),
            't1a5': Tabel_1A5.objects.all(),
            't1b': Tabel_1B_SPMI.objects.all(), 
            't2a1': Tabel_2A_Mahasiswa.objects.all().order_by('id'),
            't2a2': Tabel_2A2_Asal.objects.all().order_by('id'),
            't2a3': Tabel_2A3_Kondisi.objects.all().order_by('id'),
            't2b1': Tabel_2B1_MK.objects.all(),
            't2b2': Tabel_2B2_CPL.objects.all(),
            't2b3': Tabel_2B3_Pemenuhan.objects.all(),
            't2b4': Tabel_2B4_MasaTunggu.objects.all().order_by('id'),
            't2b5': Tabel_2B5_BidangKerja.objects.all().order_by('id'),
            't2b6': Tabel_2B6_Kepuasan.objects.all(),
            't2c': Tabel_2C_Fleksibilitas.objects.all(), 
            't2d': Tabel_2D_Rekognisi.objects.all(), 
            'sum2': Tabel_2B_Summary.objects.first(),
            't3a1': Tabel_3A1_Sarana.objects.all(),
            't3a2': Tabel_3A2_Penelitian.objects.all(),
            't3a3': Tabel_3A3_Pengembangan_DTPR.objects.all(), 
            't3c1': Tabel_3C1_Kerjasama.objects.all(),
            't3c2': Tabel_3C2_Publikasi.objects.all(),
            't3c3': Tabel_3C3_HKI.objects.all(),
            'sum3': Tabel_3_Summary.objects.first(),
            't4a1': Tabel_4A1_Sarana.objects.all(),
            't4a2': Tabel_4A2_PkM.objects.all(),
            't4c1': Tabel_4C1_Kerjasama.objects.all(),
            't4c2': Tabel_4C2_Diseminasi.objects.all(),
            't4c3': Tabel_4C3_HKI.objects.all(),
            'sum4': Tabel_4_Summary.objects.first(),
            't5_1': Tabel_5_1_TataKelola.objects.all(),
            't5_2': Tabel_5_2_Sarana.objects.all(),
            't6': Tabel_6_Misi.objects.first(),
        }
        
        if identitas_data and identitas_data.logo_pt:
            try:
                image_path = identitas_data.logo_pt.path 
                context['logo'] = InlineImage(doc, image_path, width=Mm(40))
            except Exception as e:
                print(f"Gagal memuat gambar logo: {e}")
                context['logo'] = '' 
        else:
            context['logo'] = '' 

        doc.render(context)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        response['Content-Disposition'] = 'attachment; filename="Dokumen_LKPS_Final.docx"'
        doc.save(response)
        return response

    except Exception as e:
        return HttpResponse(f"Gagal mencetak dokumen: {str(e)}", status=500)

# ==========================================
# CHATBOT & UTILS
# ==========================================
GEMINI_API_KEY = "AIzaSyBNTztqmQ978wn3w9f13r04pw3_Nd2ffvg"
genai.configure(api_key=GEMINI_API_KEY)

def chatbot_api(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        user_message = request.POST.get('message', '')
        if not user_message:
            return JsonResponse({'status': 'error', 'message': 'Pesan kosong'})

        try:
            model = genai.GenerativeModel('gemini-2.5-flash')
            system_instruction = """
            Kamu adalah "Leksi", Asisten AI ramah untuk Sistem Borang Akreditasi LKPS di Pradita University. 
            Tugasmu adalah memandu pengguna yang kebingungan saat mengisi data di aplikasi ini.
            Jawab dengan santai, sopan, dan sangat ringkas. Gunakan poin-poin.
            """
            prompt = f"{system_instruction}\n\nPertanyaan: {user_message}"
            response = model.generate_content(prompt)
            return JsonResponse({'status': 'success', 'reply': response.text})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

def db_explorer(request):
    query = request.POST.get('q', '') or request.GET.get('q', '')
    results = []
    headers = []
    error = None

    if query:
        try:
            with connection.cursor() as cursor:
                cursor.execute(query)
                if cursor.description:
                    headers = [col[0] for col in cursor.description]
                    results = cursor.fetchall()
                else:
                    results = [["Query executed successfully (no rows returned)."]]
        except Exception as e:
            error = str(e)

    table_schemas = {}
    filtered_table_names = []  
    with connection.cursor() as cursor:
        all_table_names = connection.introspection.table_names()
        for t_name in all_table_names:
            if t_name.startswith('lkps_app_'):
                filtered_table_names.append(t_name) 
                try:
                    desc = connection.introspection.get_table_description(cursor, t_name)
                    columns = [col.name for col in desc]
                    table_schemas[t_name] = columns
                except Exception:
                    table_schemas[t_name] = []

    return render(request, 'lkps_app/db_explorer.html', {
        'query': query, 'results': results, 'headers': headers, 'error': error,
        'tables': sorted(filtered_table_names), 'table_schemas_json': json.dumps(table_schemas) 
    })

def login_view(request):
    if request.method == 'POST':
        user_name = request.POST.get('username')
        pass_word = request.POST.get('password')
        user = authenticate(request, username=user_name, password=pass_word)
        if user is not None:
            auth_login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, "Username atau Password salah!")
    return render(request, 'lkps_app/login.html')

def logout_user(request):
    auth_logout(request)
    return redirect('dashboard')

def dashboard(request):
    return render(request, 'lkps_app/dashboard.html')

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
@api_view(['POST'])
def fetch_data_lppm(request):
    try:
        data_masuk = request.data
        return Response({
            "status": "success",
            "message": "Data LPPM berhasil diterima oleh sistem LKPS Universitas Pradita",
            "received_count": len(data_masuk) if isinstance(data_masuk, list) else 1
        }, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"status": "error", "message": str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==========================================
# IMPORT EXCEL TO DATABASE
# ==========================================
import openpyxl
from io import BytesIO

def import_excel(request):
    """Import data dari file Excel master ke semua tabel database."""
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        return JsonResponse({'status': 'error', 'message': 'File Excel tidak ditemukan.'}, status=400)
    
    try:
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(BytesIO(excel_file.read()), data_only=True)
        
        # Mapping: sheet_name -> (ModelClass, [list of model field names in column order])
        SHEET_MAP = {
            'Tabel_1A1': (Tabel_1A1, ['unit_kerja', 'nama_ketua', 'periode_jabatan', 'pendidikan_terakhir', 'jabatan_fungsional', 'tupoksi']),
            'Tabel_1A2_Sumber': (Tabel_1A2_Sumber, ['sumber_dana', 'ts_2', 'ts_1', 'ts', 'link_bukti']),
            'Tabel_1A3_Penggunaan': (Tabel_1A3_Penggunaan, ['penggunaan', 'ts_2', 'ts_1', 'ts', 'link_bukti']),
            'Tabel_1A4': (Tabel_1A4, ['nama_dosen', 'sks_ps_sendiri', 'sks_ps_lain', 'sks_pt_lain', 'sks_penelitian', 'sks_pkm', 'sks_tambahan']),
            'Tabel_1A5': (Tabel_1A5, ['jenis_tenaga', 's3', 's2', 's1', 'd3', 'd2_d1', 'sma_smk', 'unit_kerja']),
            'Tabel_1B_SPMI': (Tabel_1B_SPMI, ['nama_unit', 'dokumen', 'jml_auditor_cert', 'jml_auditor_non', 'frekuensi_audit', 'link_pt', 'link_upps']),
            'Tabel_2A1_Mahasiswa': (Tabel_2A_Mahasiswa, ['tahun_akademik', 'daya_tampung', 'pendaftar', 'lulus_seleksi', 'mhs_baru_reguler', 'mhs_baru_transfer', 'total_mhs_reguler', 'total_mhs_transfer']),
            'Tabel_2A2_Asal': (Tabel_2A2_Asal, ['asal_mahasiswa', 'ts_2', 'ts_1', 'ts', 'link_bukti']),
            'Tabel_2A3_Kondisi': (Tabel_2A3_Kondisi, ['status', 'ts_2', 'ts_1', 'ts']),
            'Tabel_2B1_Kurikulum': (Tabel_2B1_MK, ['nama_mk', 'sks', 'semester', 'pl1', 'pl2', 'pl3', 'pl4']),
            'Tabel_2B2_Pemetaan_CPL': (Tabel_2B2_CPL, ['kode_cpl', 'pl1', 'pl2', 'pl3', 'pl4']),
            'Tabel_2B3_Pemenuhan_CPL': (Tabel_2B3_Pemenuhan, ['cpl', 'cpmk', 'smt1', 'smt2', 'smt3']),
            'Tabel_2B4_Masa_Tunggu': (Tabel_2B4_MasaTunggu, ['tahun_lulus', 'jml_lulusan', 'jml_terlacak', 'waktu_tunggu']),
            'Tabel_2B5_Bidang_Kerja': (Tabel_2B5_BidangKerja, ['tahun_lulus', 'jml_terlacak', 'bidang_infokom', 'bidang_non_infokom', 'tingkat_multinasional', 'tingkat_nasional', 'tingkat_wirausaha']),
            'Tabel_2B6_Kepuasan': (Tabel_2B6_Kepuasan, ['jenis_kemampuan', 'sangat_baik', 'baik', 'cukup', 'kurang', 'tindak_lanjut']),
            'Tabel_2C_Fleksibilitas': (Tabel_2C_Fleksibilitas, ['bentuk_pembelajaran', 'ts_2', 'ts_1', 'ts', 'link_bukti']),
            'Tabel_2D_Rekognisi': (Tabel_2D_Rekognisi, ['sumber', 'jenis_pengakuan', 'ts_2', 'ts_1', 'ts', 'link_bukti']),
            'Tabel_3A1_Sarana': (Tabel_3A1_Sarana, ['nama_prasarana', 'daya_tampung', 'luas_ruang', 'kepemilikan', 'lisensi', 'perangkat', 'link_bukti']),
            'Tabel_3A2_Penelitian': (Tabel_3A2_Penelitian, ['nama_dtpr', 'jml_mhs', 'judul', 'jenis_hibah', 'sumber_lni', 'durasi', 'dana_ts2', 'dana_ts1', 'dana_ts', 'link_bukti']),
            'Tabel_3A3_Pengembangan': (Tabel_3A3_Pengembangan_DTPR, ['jenis_pengembangan', 'nama_dosen', 'ts_2', 'ts_1', 'ts', 'link_bukti']),
            'Tabel_3C1_Kerjasama': (Tabel_3C1_Kerjasama, ['judul', 'mitra', 'sumber_lni', 'durasi', 'dana_ts2', 'dana_ts1', 'dana_ts', 'link_bukti']),
            'Tabel_3C2_Publikasi': (Tabel_3C2_Publikasi, ['nama_dtpr', 'judul', 'jenis_pub', 'ts2', 'ts1', 'ts']),
            'Tabel_3C3_HKI': (Tabel_3C3_HKI, ['judul', 'jenis_hki', 'nama_dtpr', 'ts2', 'ts1', 'ts']),
            'Tabel_4A1_Sarana_PkM': (Tabel_4A1_Sarana, ['nama_prasarana', 'daya_tampung', 'luas_ruang', 'kepemilikan', 'lisensi', 'perangkat', 'link_bukti']),
            'Tabel_4A2_PkM': (Tabel_4A2_PkM, ['nama_dtpr', 'judul', 'jml_mhs', 'jenis_hibah', 'sumber_lni', 'durasi', 'dana_ts2', 'dana_ts1', 'dana_ts', 'link_bukti']),
            'Tabel_4C1_Kerja_PkM': (Tabel_4C1_Kerjasama, ['judul', 'mitra', 'sumber_lni', 'durasi', 'dana_ts2', 'dana_ts1', 'dana_ts', 'link_bukti']),
            'Tabel_4C2_Disem': (Tabel_4C2_Diseminasi, ['nama_dtpr', 'judul', 'lni', 'ts2', 'ts1', 'ts', 'link_bukti']),
            'Tabel_4C3_HKI_PkM': (Tabel_4C3_HKI, ['judul', 'jenis_hki', 'nama_dtpr', 'ts2', 'ts1', 'ts', 'link_bukti']),
            'Tabel_5_1_SI': (Tabel_5_1_TataKelola, ['jenis_tata_kelola', 'nama_sistem', 'akses', 'unit_pengelola', 'link_bukti']),
            'Tabel_5_2_Sarana_Pendidikan': (Tabel_5_2_Sarana, ['nama_prasarana', 'daya_tampung', 'luas_ruang', 'kepemilikan', 'lisensi', 'perangkat', 'link_bukti']),
            'Tabel_6_VisiMisi': (Tabel_6_Misi, ['visi_pt', 'visi_upps', 'visi_ps', 'misi_pt', 'misi_upps']),
        }
        
        results = {}
        
        with transaction.atomic():
            for sheet_name in wb.sheetnames:
                if sheet_name not in SHEET_MAP:
                    continue
                    
                ModelClass, field_names = SHEET_MAP[sheet_name]
                ws = wb[sheet_name]
                
                # Skip header row (row 1), read data rows
                rows_data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Skip completely empty rows
                    if not any(cell is not None and str(cell).strip() != '' for cell in row):
                        continue
                    
                    row_dict = {}
                    for col_idx, field_name in enumerate(field_names):
                        cell_value = row[col_idx] if col_idx < len(row) else None
                        
                        # Get the model field to determine type
                        try:
                            model_field = ModelClass._meta.get_field(field_name)
                        except Exception:
                            row_dict[field_name] = str(cell_value or '')
                            continue
                        
                        field_type = model_field.get_internal_type()
                        
                        if field_type in ('IntegerField',):
                            row_dict[field_name] = int(cell_value or 0)
                        elif field_type in ('DecimalField', 'FloatField'):
                            row_dict[field_name] = float(cell_value or 0)
                        elif field_type in ('BooleanField',):
                            if isinstance(cell_value, bool):
                                row_dict[field_name] = cell_value
                            else:
                                row_dict[field_name] = str(cell_value or '').strip().lower() in ('1', 'true', 'ya', 'yes', 'v', '✓')
                        elif field_type in ('URLField',):
                            row_dict[field_name] = str(cell_value or '') if cell_value else ''
                        elif field_type in ('TextField',):
                            row_dict[field_name] = str(cell_value or '')
                        else:
                            row_dict[field_name] = str(cell_value or '')
                    
                    rows_data.append(ModelClass(**row_dict))
                
                # Only replace data if the sheet has content
                if rows_data:
                    ModelClass.objects.all().delete()
                    ModelClass.objects.bulk_create(rows_data)
                    results[sheet_name] = len(rows_data)
        
        return JsonResponse({
            'status': 'success',
            'message': f'Berhasil mengimpor {sum(results.values())} baris dari {len(results)} tabel.',
            'details': results
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Gagal mengimpor: {str(e)}'}, status=500)

# ==========================================
# UNIVERSAL AUTOSAVE HELPER
# ==========================================
def universal_table_autosave(request, ModelClass, field_mapping):
    """
    Fungsi master untuk memproses autosave multi-baris secara otomatis.
    """
    try:
        with transaction.atomic():
            ModelClass.objects.all().delete()
            html_fields = list(field_mapping.keys())
            if not html_fields:
                return JsonResponse({'status': 'error', 'message': 'Mapping kosong'})
                
            # Cek panjang data dari input pertama
            base_data = request.POST.getlist(html_fields[0])
            jumlah_baris = len(base_data)
            data_yang_akan_disimpan = []
            
            for i in range(jumlah_baris):
                row_data = {}
                baris_ada_isinya = False
                
                for html_name, model_field_info in field_mapping.items():
                    model_field = model_field_info['field']
                    tipe_data = model_field_info['type']
                    
                    nilai_list = request.POST.getlist(html_name)
                    nilai_mentah = nilai_list[i] if i < len(nilai_list) else ""
                    
                    if tipe_data == 'int':
                        nilai_bersih = int(nilai_mentah or 0)
                        if nilai_bersih > 0 or nilai_mentah != "": baris_ada_isinya = True
                    elif tipe_data == 'float':
                        nilai_bersih = float(nilai_mentah or 0.0)
                        if nilai_bersih > 0 or nilai_mentah != "": baris_ada_isinya = True
                    elif tipe_data == 'bool':
                        nilai_bersih = (nilai_mentah == '1')
                        if nilai_bersih: baris_ada_isinya = True
                    else: 
                        nilai_bersih = nilai_mentah.strip()
                        if nilai_bersih: baris_ada_isinya = True
                        
                    row_data[model_field] = nilai_bersih
                
                if baris_ada_isinya:
                    data_yang_akan_disimpan.append(ModelClass(**row_data))
            
            ModelClass.objects.bulk_create(data_yang_akan_disimpan)
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# ==========================================
# IDENTITAS & PROGRAM STUDI
# ==========================================
def sampul(request):
    prodi_default = ProgramStudi.objects.first()
    if not prodi_default:
        prodi_default = ProgramStudi.objects.create(nama_prodi="Prodi Informatika", jenjang_studi="S1", akreditasi="Baik", no_sk="-")

    identitas, created = IdentitasPengusul.objects.get_or_create(id=1, defaults={'program_studi': prodi_default})

    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        mapping = {
            'sampul_nama_ps': 'nama_ps_sampul', 'sampul_nama_pt': 'nama_pt_sampul',
            'sampul_kota': 'kota_sampul', 'sampul_tahun': 'tahun_sampul',
            'pengusul_pt': 'perguruan_tinggi', 'pengusul_upps': 'unit_pengelola',
            'pengusul_alamat': 'alamat', 'pengusul_telepon': 'telepon',
            'pengusul_email_web': 'email_web', 'pengusul_sk_pt': 'sk_pt', 
            'pengusul_tgl_sk_pt': 'tgl_sk_pt', 'pengusul_sk_ps': 'sk_ps', 'pengusul_tgl_sk_ps': 'tgl_sk_ps',
        }

        for html_name, model_field in mapping.items():
            if html_name in request.POST:
                val = request.POST.get(html_name)
                date_and_int_fields = ['tahun_sampul', 'tgl_sk_pt', 'tgl_sk_ps', 'tahun_mhs_pertama']
                if model_field in date_and_int_fields and val == "": val = None
                elif val is None: val = ""
                setattr(identitas, model_field, val)
        
        if 'logo_pt' in request.FILES:
            identitas.logo_pt = request.FILES['logo_pt']
            
        identitas.save() 

        # Logika Tim Penyusun
        namas = request.POST.getlist('penyusun_nama[]')
        nidns = request.POST.getlist('penyusun_nidn[]')
        jabatans = request.POST.getlist('penyusun_jabatan[]')
        tanggals = request.POST.getlist('penyusun_tanggal[]')

        identitas.tim_penyusun.all().delete()
        for i in range(len(namas)):
            nama_clean = namas[i].strip() if namas[i] else ""
            if nama_clean: 
                TimPenyusun.objects.create(
                    identitas=identitas, nama=nama_clean,
                    nidn=nidns[i] if i < len(nidns) else "",
                    jabatan=jabatans[i] if i < len(jabatans) else "",
                    tanggal_pengisian=tanggals[i] if (i < len(tanggals) and tanggals[i]) else None
                )
        return JsonResponse({'status': 'success'})

    return render(request, 'lkps_app/sampul.html', {'data': identitas, 'tim_penyusun': identitas.tim_penyusun.all()})

def program_studi(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            ProgramStudi.objects.create(
                nama_prodi = request.POST.get('nama_prodi'),
                jenjang_studi = request.POST.get('jenjang_prodi'),
                akreditasi = request.POST.get('status_akreditasi'),
                no_sk = request.POST.get('sk_banpt')       
            )
            return JsonResponse({'status': 'success', 'message': 'Data berhasil disimpan!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return render(request, 'lkps_app/program_studi.html', {'daftar_prodi': ProgramStudi.objects.all().order_by('id')})

def manajemen_user(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            input_email = request.POST.get('email')
            user_baru = User.objects.create_user(
                username=input_email, email=input_email,
                password=request.POST.get('password'), first_name=request.POST.get('nama_lengkap') 
            )
            ProfilPengguna.objects.create(
                user=user_baru, nomor_induk=request.POST.get('nomor_induk'),
                role=request.POST.get('role'), akses_prodi=request.POST.get('akses_prodi')
            )
            return JsonResponse({'status': 'success', 'message': 'Akun berhasil dibuat!'})
        except IntegrityError:
            return JsonResponse({'status': 'error', 'message': f'Email {input_email} sudah digunakan!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    daftar_pengguna = []
    for profil in ProfilPengguna.objects.select_related('user').all():
        daftar_pengguna.append({
            'nama_lengkap': profil.user.first_name, 'email': profil.user.email,
            'nomor_induk': profil.nomor_induk, 'role': profil.role,
            'akses_prodi': profil.akses_prodi, 'is_active': profil.user.is_active
        })
    return render(request, 'lkps_app/manajemen_user.html', {'daftar_pengguna': daftar_pengguna})


# ==========================================
# KRITERIA 1: VIEWS DENGAN LOGIKA KHUSUS
# ==========================================
def tabel_1a_dana(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            with transaction.atomic():
                # Sumber
                mapping_sumber = {
                    'sumber_dana[]': {'field': 'sumber_dana', 'type': 'str'},
                    'sumber_ts2[]': {'field': 'ts_2', 'type': 'float'},
                    'sumber_ts1[]': {'field': 'ts_1', 'type': 'float'},
                    'sumber_ts[]': {'field': 'ts', 'type': 'float'},
                    'link_sumber[]': {'field': 'link_bukti', 'type': 'str'}
                }
                res1 = universal_table_autosave(request, Tabel_1A2_Sumber, mapping_sumber)
                if json.loads(res1.content).get('status') == 'error': raise Exception(json.loads(res1.content).get('message'))
                
                # Penggunaan
                mapping_guna = {
                    'guna_dana[]': {'field': 'penggunaan', 'type': 'str'},
                    'guna_ts2[]': {'field': 'ts_2', 'type': 'float'},
                    'guna_ts1[]': {'field': 'ts_1', 'type': 'float'},
                    'guna_ts[]': {'field': 'ts', 'type': 'float'},
                    'link_guna[]': {'field': 'link_bukti', 'type': 'str'}
                }
                res2 = universal_table_autosave(request, Tabel_1A3_Penggunaan, mapping_guna)
                if json.loads(res2.content).get('status') == 'error': raise Exception(json.loads(res2.content).get('message'))
                
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    ctx = {'data_sumber': Tabel_1A2_Sumber.objects.all().order_by('id'), 'data_penggunaan': Tabel_1A3_Penggunaan.objects.all().order_by('id')}
    return render(request, 'lkps_app/tabel_1a_dana.html', ctx)

def tabel_1a4(request):
    # Tabel ini agak unik karena ada penggabungan field tambahan, jadi kita pertahankan manual
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        nama = request.POST.getlist('nama_dtpr[]')
        ps_s = request.POST.getlist('sks_pengajaran_ps_sendiri[]')
        ps_l = request.POST.getlist('sks_pengajaran_ps_lain[]')
        pt_l = request.POST.getlist('sks_pengajaran_pt_lain[]')
        pen = request.POST.getlist('sks_penelitian[]')
        pkm = request.POST.getlist('sks_pengabdian[]')
        m_s = request.POST.getlist('sks_manajemen_pt_sendiri[]')
        m_l = request.POST.getlist('sks_manajemen_pt_lain[]')

        try:
            with transaction.atomic():
                Tabel_1A4.objects.all().delete()
                for i in range(len(nama)):
                    if nama[i].strip():
                        Tabel_1A4.objects.create(
                            nama_dosen=nama[i],
                            sks_ps_sendiri=float(ps_s[i] or 0), sks_ps_lain=float(ps_l[i] or 0), sks_pt_lain=float(pt_l[i] or 0),
                            sks_penelitian=float(pen[i] or 0), sks_pkm=float(pkm[i] or 0),
                            sks_manajemen_pt_sendiri = float(m_s[i] or 0),
                            sks_manajemen_pt_lain = float(m_l[i] or 0)
                        )
            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})
    return render(request, 'lkps_app/tabel_1a4.html', {'data': Tabel_1A4.objects.all()})

def tabel_1a5(request):
    # Sama, ada penggabungan (D2+D1, S1+D4)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        jenis = request.POST.getlist('jenis_tendik[]')
        s3 = request.POST.getlist('jml_s3[]'); s2 = request.POST.getlist('jml_s2[]'); s1 = request.POST.getlist('jml_s1[]')
        d4 = request.POST.getlist('jml_d4[]'); d3 = request.POST.getlist('jml_d3[]'); d2 = request.POST.getlist('jml_d2[]')
        d1 = request.POST.getlist('jml_d1[]'); sma = request.POST.getlist('jml_sma[]'); unit = request.POST.getlist('unit_kerja[]')
        try:
            with transaction.atomic():
                Tabel_1A5.objects.all().delete()
                for i in range(len(jenis)):
                    if jenis[i].strip():
                        Tabel_1A5.objects.create(
                            jenis_tenaga=jenis[i], s3=int(s3[i] or 0), s2=int(s2[i] or 0),
                            s1=int(s1[i] or 0) + int(d4[i] or 0), d3=int(d3[i] or 0),
                            d2_d1=int(d2[i] or 0) + int(d1[i] or 0), sma_smk=int(sma[i] or 0), unit_kerja=unit[i]
                        )
            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})
    return render(request, 'lkps_app/tabel_1a5.html', {'data': Tabel_1A5.objects.all()})

# ==========================================
# KRITERIA 2: VIEWS DENGAN LOGIKA KHUSUS
# ==========================================
def tabel_2a1(request):
    for y in ['TS-3', 'TS-2', 'TS-1', 'TS']: Tabel_2A_Mahasiswa.objects.get_or_create(tahun_akademik=y)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            with transaction.atomic():
                thn = request.POST.getlist('tahun_akademik[]'); daya = request.POST.getlist('daya_tampung[]')
                p_reg = request.POST.getlist('pendaftar_reg[]'); p_afi = request.POST.getlist('pendaftar_afi[]'); p_khu = request.POST.getlist('pendaftar_khu[]')
                l_reg = request.POST.getlist('lulus_reg[]'); l_afi = request.POST.getlist('lulus_afi[]'); l_khu = request.POST.getlist('lulus_khu[]'); l_rpl = request.POST.getlist('lulus_rpl[]')
                m_reg = request.POST.getlist('maba_reg[]'); m_afi = request.POST.getlist('maba_afi[]'); m_khu = request.POST.getlist('maba_khu[]'); m_rpl = request.POST.getlist('maba_rpl[]')
                a_reg = request.POST.getlist('aktif_reg[]'); a_afi = request.POST.getlist('aktif_afi[]'); a_khu = request.POST.getlist('aktif_khu[]'); a_rpl = request.POST.getlist('aktif_rpl[]')
                
                for i in range(len(thn)):
                    obj = Tabel_2A_Mahasiswa.objects.get(tahun_akademik=thn[i])
                    obj.daya_tampung = int(daya[i] or 0)
                    obj.pendaftar = int(p_reg[i] or 0) + int(p_afi[i] or 0) + int(p_khu[i] or 0)
                    obj.lulus_seleksi = int(l_reg[i] or 0) + int(l_afi[i] or 0) + int(l_khu[i] or 0) + int(l_rpl[i] or 0)
                    obj.mhs_baru_reguler = int(m_reg[i] or 0) + int(m_afi[i] or 0) + int(m_khu[i] or 0)
                    obj.mhs_baru_transfer = int(m_rpl[i] or 0)
                    obj.total_mhs_reguler = int(a_reg[i] or 0) + int(a_afi[i] or 0) + int(a_khu[i] or 0)
                    obj.total_mhs_transfer = int(a_rpl[i] or 0)
                    obj.save()
            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})
    return render(request, 'lkps_app/tabel_2a1.html', {'data_2a1': Tabel_2A_Mahasiswa.objects.all().order_by('id')})

def tabel_2a2(request):
    kategori_asal = ["Kota/Kab sama dengan PS", "Kota/Kabupaten Lain", "Provinsi Lain", "Negara Lain", "Afirmasi", "Berkebutuhan Khusus"]
    for asal in kategori_asal: Tabel_2A2_Asal.objects.get_or_create(asal_mahasiswa=asal)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            with transaction.atomic():
                kat_asal = request.POST.getlist('asal_kategori[]')
                asal_ts2 = request.POST.getlist('asal_ts2[]'); asal_ts1 = request.POST.getlist('asal_ts1[]'); asal_ts = request.POST.getlist('asal_ts[]')
                link_asal = request.POST.getlist('link_asal[]')
                for i in range(len(kat_asal)):
                    obj_asal = Tabel_2A2_Asal.objects.get(asal_mahasiswa=kat_asal[i])
                    obj_asal.ts_2 = int(asal_ts2[i] or 0); obj_asal.ts_1 = int(asal_ts1[i] or 0); obj_asal.ts = int(asal_ts[i] or 0)
                    obj_asal.link_bukti = link_asal[i]
                    obj_asal.save()
            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})
    return render(request, 'lkps_app/tabel_2a2.html', {'data_2a2': Tabel_2A2_Asal.objects.all().order_by('id')})

def tabel_2a3(request):
    kategori_kondisi = ["Mahasiswa Aktif pada saat TS", "Lulus pada saat TS", "Mengundurkan Diri/DO pada saat TS"]
    for kondisi in kategori_kondisi: Tabel_2A3_Kondisi.objects.get_or_create(status=kondisi)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            with transaction.atomic():
                kat_kon = request.POST.getlist('kondisi_kategori[]')
                kon_ts2 = request.POST.getlist('kondisi_ts2[]'); kon_ts1 = request.POST.getlist('kondisi_ts1[]'); kon_ts = request.POST.getlist('kondisi_ts[]')
                for i in range(len(kat_kon)):
                    obj_kon = Tabel_2A3_Kondisi.objects.get(status=kat_kon[i])
                    obj_kon.ts_2 = int(kon_ts2[i] or 0); obj_kon.ts_1 = int(kon_ts1[i] or 0); obj_kon.ts = int(kon_ts[i] or 0)
                    obj_kon.save()
            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})
    return render(request, 'lkps_app/tabel_2a3.html', {'data_2a3': Tabel_2A3_Kondisi.objects.all().order_by('id')})

def tabel_2a4(request):
    years = ['TS-2', 'TS-1', 'TS']
    for y in years: Tabel_2B4_MasaTunggu.objects.get_or_create(tahun_lulus=y)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            with transaction.atomic():
                thn_mt = request.POST.getlist('tahun_lulus_mt[]')
                j_lul_mt = request.POST.getlist('jml_lulusan_mt[]')
                j_ter_mt = request.POST.getlist('jml_terlacak_mt[]')
                w_tng = request.POST.getlist('waktu_tunggu[]')
                for i in range(len(thn_mt)):
                    obj = Tabel_2B4_MasaTunggu.objects.get(tahun_lulus=thn_mt[i])
                    obj.jml_lulusan = int(j_lul_mt[i] or 0); obj.jml_terlacak = int(j_ter_mt[i] or 0); obj.waktu_tunggu = float(w_tng[i] or 0)
                    obj.save()
            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})
    return render(request, 'lkps_app/tabel_2a4.html', {'data_2b4': Tabel_2B4_MasaTunggu.objects.all().order_by('id')})

def tabel_2a5(request):
    years = ['TS-2', 'TS-1', 'TS']
    for y in years: Tabel_2B5_BidangKerja.objects.get_or_create(tahun_lulus=y)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            with transaction.atomic():
                thn_bk = request.POST.getlist('tahun_lulus_bk[]')
                j_lul_bk = request.POST.getlist('jml_lulusan_bk[]'); j_ter_bk = request.POST.getlist('jml_terlacak_bk[]')
                b_info = request.POST.getlist('bidang_infokom[]'); b_non = request.POST.getlist('bidang_non_infokom[]')
                t_multi = request.POST.getlist('tingkat_multinasional[]'); t_nas = request.POST.getlist('tingkat_nasional[]'); t_wira = request.POST.getlist('tingkat_wirausaha[]')
                for i in range(len(thn_bk)):
                    obj = Tabel_2B5_BidangKerja.objects.get(tahun_lulus=thn_bk[i])
                    obj.jml_lulusan = int(j_lul_bk[i] or 0); obj.jml_terlacak = int(j_ter_bk[i] or 0)
                    obj.bidang_infokom = int(b_info[i] or 0); obj.bidang_non_infokom = int(b_non[i] or 0)
                    obj.tingkat_multinasional = int(t_multi[i] or 0); obj.tingkat_nasional = int(t_nas[i] or 0); obj.tingkat_wirausaha = int(t_wira[i] or 0)
                    obj.save()
            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})
    return render(request, 'lkps_app/tabel_2a5.html', {'data_2b5': Tabel_2B5_BidangKerja.objects.all().order_by('id')})

def tabel_2a6(request):
    kemampuans = ["Kerjasama Tim", "Keahlian di Bidang Prodi", "Kemampuan Berbahasa Asing (Inggris)", "Kemampuan Berkomunikasi", "Pengembangan Diri", "Kepemimpinan", "Etos Kerja"]
    for k in kemampuans: Tabel_2B6_Kepuasan.objects.get_or_create(jenis_kemampuan=k)
    Tabel_2B_Summary.objects.get_or_create(id=1)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            with transaction.atomic():
                jenis_kpu = request.POST.getlist('jenis_kemampuan[]')
                sb = request.POST.getlist('sangat_baik[]'); b = request.POST.getlist('baik[]'); c = request.POST.getlist('cukup[]'); k = request.POST.getlist('kurang[]')
                tl = request.POST.getlist('tindak_lanjut[]')
                for i in range(len(jenis_kpu)):
                    obj = Tabel_2B6_Kepuasan.objects.get(jenis_kemampuan=jenis_kpu[i])
                    obj.sangat_baik = float(sb[i] or 0); obj.baik = float(b[i] or 0); obj.cukup = float(c[i] or 0); obj.kurang = float(k[i] or 0); obj.tindak_lanjut = tl[i]
                    obj.save()
                sum_obj = Tabel_2B_Summary.objects.get(id=1)
                sum_obj.total_alumni_3thn = int(request.POST.get('total_alumni_3thn') or 0)
                sum_obj.total_responden = int(request.POST.get('total_responden') or 0)
                sum_obj.total_mhs_aktif_ts = int(request.POST.get('total_mhs_aktif_ts') or 0)
                sum_obj.save()
            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})
    return render(request, 'lkps_app/tabel_2a6.html', {'data_2b6': Tabel_2B6_Kepuasan.objects.all().order_by('id'), 'summary': Tabel_2B_Summary.objects.get(id=1)})

# ==========================================
# KRITERIA 3, 4, 5 & 6: VIEWS DENGAN LOGIKA KHUSUS
# ==========================================
def tabel_3_penelitian(request):
    Tabel_3_Summary.objects.get_or_create(id=1)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            with transaction.atomic():
                # 3A1
                mapping_3a1 = {
                    'nama_prasarana[]': {'field': 'nama_prasarana', 'type': 'str'},
                    'daya_tampung[]': {'field': 'daya_tampung', 'type': 'int'},
                    'luas_ruang[]': {'field': 'luas_ruang', 'type': 'float'},
                    'kepemilikan[]': {'field': 'kepemilikan', 'type': 'str'},
                    'lisensi[]': {'field': 'lisensi', 'type': 'str'},
                    'perangkat[]': {'field': 'perangkat', 'type': 'str'},
                    'link_bukti_prasarana[]': {'field': 'link_bukti', 'type': 'str'}
                }
                universal_table_autosave(request, Tabel_3A1_Sarana, mapping_3a1)
                
                # 3A2
                mapping_3a2 = {
                    'nama_dtpr_ketua[]': {'field': 'nama_dtpr', 'type': 'str'},
                    'jml_mhs_terlibat[]': {'field': 'jml_mhs', 'type': 'int'},
                    'judul_penelitian[]': {'field': 'judul', 'type': 'str'},
                    'jenis_hibah[]': {'field': 'jenis_hibah', 'type': 'str'},
                    'sumber_lni[]': {'field': 'sumber_lni', 'type': 'str'},
                    'durasi_tahun[]': {'field': 'durasi', 'type': 'float'},
                    'dana_ts2[]': {'field': 'dana_ts2', 'type': 'float'},
                    'dana_ts1[]': {'field': 'dana_ts1', 'type': 'float'},
                    'dana_ts[]': {'field': 'dana_ts', 'type': 'float'},
                    'link_bukti_penelitian[]': {'field': 'link_bukti', 'type': 'str'}
                }
                universal_table_autosave(request, Tabel_3A2_Penelitian, mapping_3a2)

                # 3C1
                mapping_3c1 = {
                    'judul_kerjasama[]': {'field': 'judul', 'type': 'str'},
                    'mitra_kerjasama[]': {'field': 'mitra', 'type': 'str'},
                    'sumber_kerjasama_lni[]': {'field': 'sumber_lni', 'type': 'str'},
                    'durasi_kerjasama[]': {'field': 'durasi', 'type': 'float'},
                    'dana_k_ts2[]': {'field': 'dana_ts2', 'type': 'float'},
                    'dana_k_ts1[]': {'field': 'dana_ts1', 'type': 'float'},
                    'dana_k_ts[]': {'field': 'dana_ts', 'type': 'float'},
                    'link_bukti_kerjasama[]': {'field': 'link_bukti', 'type': 'str'}
                }
                universal_table_autosave(request, Tabel_3C1_Kerjasama, mapping_3c1)

                # 3C2
                mapping_3c2 = {
                    'nama_dtpr_pub[]': {'field': 'nama_dtpr', 'type': 'str'},
                    'judul_pub[]': {'field': 'judul', 'type': 'str'},
                    'jenis_pub[]': {'field': 'jenis_pub', 'type': 'str'},
                    'pub_ts2[]': {'field': 'ts2', 'type': 'bool'},
                    'pub_ts1[]': {'field': 'ts1', 'type': 'bool'},
                    'pub_ts[]': {'field': 'ts', 'type': 'bool'}
                }
                universal_table_autosave(request, Tabel_3C2_Publikasi, mapping_3c2)

                # 3C3
                mapping_3c3 = {
                    'judul_hki[]': {'field': 'judul', 'type': 'str'},
                    'jenis_hki[]': {'field': 'jenis_hki', 'type': 'str'},
                    'nama_dtpr_hki[]': {'field': 'nama_dtpr', 'type': 'str'},
                    'hki_ts2[]': {'field': 'ts2', 'type': 'bool'},
                    'hki_ts1[]': {'field': 'ts1', 'type': 'bool'},
                    'hki_ts[]': {'field': 'ts', 'type': 'bool'}
                }
                universal_table_autosave(request, Tabel_3C3_HKI, mapping_3c3)

                # Summary
                summary = Tabel_3_Summary.objects.get(id=1)
                summary.link_roadmap = request.POST.get('link_roadmap_penelitian')
                summary.total_jenis_hibah = int(request.POST.get('total_jenis_hibah') or 0)
                summary.total_mitra = int(request.POST.get('total_mitra') or 0)
                summary.save()

            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})

    ctx = {
        'data_3a1': Tabel_3A1_Sarana.objects.all(), 'data_3a2': Tabel_3A2_Penelitian.objects.all(),
        'data_3c1': Tabel_3C1_Kerjasama.objects.all(), 'data_3c2': Tabel_3C2_Publikasi.objects.all(),
        'data_3c3': Tabel_3C3_HKI.objects.all(), 'summary': Tabel_3_Summary.objects.get(id=1),
    }
    return render(request, 'lkps_app/tabel_3_penelitian.html', ctx)

def tabel_4_pkm(request):
    Tabel_4_Summary.objects.get_or_create(id=1)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            with transaction.atomic():
                # 4A1
                mapping_4a1 = {
                    'nama_prasarana_pkm[]': {'field': 'nama_prasarana', 'type': 'str'},
                    'daya_tampung_pkm[]': {'field': 'daya_tampung', 'type': 'int'},
                    'luas_ruang_pkm[]': {'field': 'luas_ruang', 'type': 'float'},
                    'kepemilikan_pkm[]': {'field': 'kepemilikan', 'type': 'str'},
                    'lisensi_pkm[]': {'field': 'lisensi', 'type': 'str'},
                    'perangkat_pkm[]': {'field': 'perangkat', 'type': 'str'},
                    'link_bukti_prasarana_pkm[]': {'field': 'link_bukti', 'type': 'str'}
                }
                universal_table_autosave(request, Tabel_4A1_Sarana, mapping_4a1)

                # 4A2
                mapping_4a2 = {
                    'nama_dtpr_pkm[]': {'field': 'nama_dtpr', 'type': 'str'},
                    'judul_kegiatan_pkm[]': {'field': 'judul', 'type': 'str'},
                    'jml_mhs_pkm[]': {'field': 'jml_mhs', 'type': 'int'},
                    'jenis_hibah_pkm[]': {'field': 'jenis_hibah', 'type': 'str'},
                    'sumber_dana_pkm[]': {'field': 'sumber_lni', 'type': 'str'},
                    'durasi_pkm[]': {'field': 'durasi', 'type': 'float'},
                    'dana_pkm_ts2[]': {'field': 'dana_ts2', 'type': 'float'},
                    'dana_pkm_ts1[]': {'field': 'dana_ts1', 'type': 'float'},
                    'dana_pkm_ts[]': {'field': 'dana_ts', 'type': 'float'},
                    'link_bukti_dana_pkm[]': {'field': 'link_bukti', 'type': 'str'}
                }
                universal_table_autosave(request, Tabel_4A2_PkM, mapping_4a2)

                # 4C1
                mapping_4c1 = {
                    'judul_kerjasama_pkm[]': {'field': 'judul', 'type': 'str'},
                    'mitra_kerjasama_pkm[]': {'field': 'mitra', 'type': 'str'},
                    'sumber_kerja_pkm[]': {'field': 'sumber_lni', 'type': 'str'},
                    'durasi_kerja_pkm[]': {'field': 'durasi', 'type': 'float'},
                    'dana_kpkm_ts2[]': {'field': 'dana_ts2', 'type': 'float'},
                    'dana_kpkm_ts1[]': {'field': 'dana_ts1', 'type': 'float'},
                    'dana_kpkm_ts[]': {'field': 'dana_ts', 'type': 'float'},
                    'link_bukti_kerja_pkm[]': {'field': 'link_bukti', 'type': 'str'}
                }
                universal_table_autosave(request, Tabel_4C1_Kerjasama, mapping_4c1)

                # 4C2
                mapping_4c2 = {
                    'nama_dtpr_disem[]': {'field': 'nama_dtpr', 'type': 'str'},
                    'judul_disem[]': {'field': 'judul', 'type': 'str'},
                    'lni_disem[]': {'field': 'lni', 'type': 'str'},
                    'disem_ts2[]': {'field': 'ts2', 'type': 'bool'},
                    'disem_ts1[]': {'field': 'ts1', 'type': 'bool'},
                    'disem_ts[]': {'field': 'ts', 'type': 'bool'},
                    'link_bukti_disem[]': {'field': 'link_bukti', 'type': 'str'}
                }
                universal_table_autosave(request, Tabel_4C2_Diseminasi, mapping_4c2)

                # 4C3
                mapping_4c3 = {
                    'judul_hki_pkm[]': {'field': 'judul', 'type': 'str'},
                    'jenis_hki_pkm[]': {'field': 'jenis_hki', 'type': 'str'},
                    'nama_dtpr_hki_pkm[]': {'field': 'nama_dtpr', 'type': 'str'},
                    'hkipkm_ts2[]': {'field': 'ts2', 'type': 'bool'},
                    'hkipkm_ts1[]': {'field': 'ts1', 'type': 'bool'},
                    'hkipkm_ts[]': {'field': 'ts', 'type': 'bool'},
                    'link_bukti_hki_pkm[]': {'field': 'link_bukti', 'type': 'str'}
                }
                universal_table_autosave(request, Tabel_4C3_HKI, mapping_4c3)

                # Summary
                summary = Tabel_4_Summary.objects.get(id=1)
                summary.link_roadmap = request.POST.get('link_roadmap_pkm')
                summary.total_jenis_hibah = int(request.POST.get('total_jenis_hibah_pkm') or 0)
                summary.jml_disem_hasil = int(request.POST.get('jml_disem_hasil') or 0)
                summary.save()

            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})

    ctx = {
        'data_4a1': Tabel_4A1_Sarana.objects.all(), 'data_4a2': Tabel_4A2_PkM.objects.all(),
        'data_4c1': Tabel_4C1_Kerjasama.objects.all(), 'data_4c2': Tabel_4C2_Diseminasi.objects.all(),
        'data_4c3': Tabel_4C3_HKI.objects.all(), 'summary': Tabel_4_Summary.objects.get(id=1),
    }
    return render(request, 'lkps_app/tabel_4_pkm.html', ctx)

def tabel_5_akuntabilitas(request):
    jenis_default = ["Pendidikan", "Keuangan", "SDM", "Sarana Prasarana", "Sistem Penjaminan Mutu"]
    for jenis in jenis_default: Tabel_5_1_TataKelola.objects.get_or_create(jenis_tata_kelola=jenis)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            with transaction.atomic():
                mapping_51 = {
                    'jenis_tata_kelola[]': {'field': 'jenis_tata_kelola', 'type': 'str'},
                    'nama_sistem_info[]': {'field': 'nama_sistem', 'type': 'str'},
                    'akses_sistem[]': {'field': 'akses', 'type': 'str'},
                    'unit_pengelola_sistem[]': {'field': 'unit_pengelola', 'type': 'str'},
                    'link_bukti_sistem[]': {'field': 'link_bukti', 'type': 'str'}
                }
                universal_table_autosave(request, Tabel_5_1_TataKelola, mapping_51)

                mapping_52 = {
                    'nama_prasarana_pend[]': {'field': 'nama_prasarana', 'type': 'str'},
                    'daya_tampung_pend[]': {'field': 'daya_tampung', 'type': 'int'},
                    'luas_ruang_pend[]': {'field': 'luas_ruang', 'type': 'float'},
                    'kepemilikan_pend[]': {'field': 'kepemilikan', 'type': 'str'},
                    'lisensi_pend[]': {'field': 'lisensi', 'type': 'str'},
                    'perangkat_pend[]': {'field': 'perangkat', 'type': 'str'},
                    'link_bukti_prasarana_pend[]': {'field': 'link_bukti', 'type': 'str'}
                }
                universal_table_autosave(request, Tabel_5_2_Sarana, mapping_52)
            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})

    ctx = {'data_5_1': Tabel_5_1_TataKelola.objects.all(), 'data_5_2': Tabel_5_2_Sarana.objects.all()}
    return render(request, 'lkps_app/tabel_5_akuntabilitas.html', ctx)

def tabel_6_misi(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            Tabel_6_Misi.objects.update_or_create(
                id=1, 
                defaults={
                    'visi_pt': request.POST.get('visi_pt'), 'visi_upps': request.POST.get('visi_upps'),
                    'visi_ps': request.POST.get('visi_ps'), 'misi_pt': request.POST.get('misi_pt'),
                    'misi_upps': request.POST.get('misi_upps')
                }
            )
            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)})
    data, created = Tabel_6_Misi.objects.get_or_create(id=1)
    return render(request, 'lkps_app/tabel_6_misi.html', {'data': data})